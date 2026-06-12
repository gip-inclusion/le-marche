import logging
import re
from decimal import Decimal, InvalidOperation

import openpyxl

from lemarche.siaes.models import Siae


logger = logging.getLogger(__name__)

# Mapping from normalized header variants → canonical column name
_HEADER_ALIASES = {
    "raison sociale": "supplier_name",
    "raison social": "supplier_name",
    "raison sociale du fournisseur": "supplier_name",
    "fournisseur": "supplier_name",
    "nom du fournisseur": "supplier_name",
    "siret": "supplier_siret",
    "siret du fournisseur": "supplier_siret",
    "depense": "purchase_amount",
    "dépense": "purchase_amount",
    "depense achat": "purchase_amount",
    "dépense achat": "purchase_amount",
    "montant": "purchase_amount",
    "montant ht": "purchase_amount",
    "budget": "purchase_amount",
    "categorie d'achat": "purchase_category",
    "catégorie d'achat": "purchase_category",
    "categorie": "purchase_category",
    "catégorie": "purchase_category",
    "entite acheteuse": "purchase_entity",
    "entité acheteuse": "purchase_entity",
    "entite": "purchase_entity",
    "entité": "purchase_entity",
}

REQUIRED_COLUMNS = {"supplier_name", "supplier_siret", "purchase_amount"}


def _normalize(text: str) -> str:
    """Lowercase, strip, remove parenthetical suffixes like '(€)' or '(optionnelle)'."""
    text = text.lower().strip()
    text = re.sub(r"\s*\(.*?\)", "", text).strip()
    return text


def _detect_columns(header_row) -> dict[str, int]:
    """Returns {canonical_name: col_index} for recognized columns."""
    mapping = {}
    for idx, cell in enumerate(header_row):
        if cell.value is None:
            continue
        normalized = _normalize(str(cell.value))
        canonical = _HEADER_ALIASES.get(normalized)
        if canonical:
            mapping[canonical] = idx
    return mapping


def parse_purchases_excel(file_obj, year: int, company) -> tuple[list, int, int]:
    """
    Parse an Excel file of purchase data and create Purchase objects.

    Returns (errors, total_rows, matched_rows).
    Raises ValueError for unrecoverable format errors.
    """
    from lemarche.purchases.models import Purchase

    try:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    except Exception:
        raise ValueError("Le fichier n'est pas un fichier Excel valide (.xlsx).")

    ws = wb.active
    rows = list(ws.iter_rows())

    if not rows:
        raise ValueError("Le fichier est vide.")

    # Find header row among first 10 rows
    col_map = {}
    header_row_idx = None
    for i, row in enumerate(rows[:10]):
        col_map = _detect_columns(row)
        if REQUIRED_COLUMNS.issubset(col_map.keys()):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError(
            "En-têtes manquants. Colonnes attendues : "
            "« Raison sociale du Fournisseur », « SIRET », « Dépense achat (€) »."
        )

    data_rows = rows[header_row_idx + 1 :]
    if not data_rows:
        raise ValueError("Le fichier ne contient aucune ligne de données.")

    # Pre-load live SIAEs indexed by SIRET for fast lookup
    siae_by_siret: dict[str, Siae] = {s.siret: s for s in Siae.objects.is_live().only("id", "siret")}

    purchases = []
    errors = []
    matched_rows = 0

    for row_num, row in enumerate(data_rows, start=header_row_idx + 2):
        cells = [cell.value for cell in row]

        def get(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(cells):
                return None
            val = cells[idx]
            return str(val).strip() if val is not None else None

        supplier_name = get("supplier_name") or ""
        supplier_siret = (get("supplier_siret") or "").replace(" ", "").replace("-", "")
        amount_raw = get("purchase_amount") or ""
        purchase_category = get("purchase_category") or None
        buying_entity = get("purchase_entity") or None

        # Skip blank rows
        if not supplier_name and not supplier_siret:
            continue

        if not supplier_name:
            errors.append(f"Ligne {row_num} : raison sociale manquante.")
            continue

        if not supplier_siret or not supplier_siret.isdigit() or len(supplier_siret) != 14:
            errors.append(f"Ligne {row_num} : SIRET invalide « {supplier_siret} » (14 chiffres attendus).")
            continue

        try:
            amount_str = re.sub(r"[^\d,.\-]", "", str(amount_raw)).replace(",", ".")
            purchase_amount = Decimal(amount_str)
            if purchase_amount <= 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            errors.append(f"Ligne {row_num} : montant invalide « {amount_raw} ».")
            continue

        siae = siae_by_siret.get(supplier_siret)
        if siae:
            matched_rows += 1

        purchases.append(
            Purchase(
                supplier_name=supplier_name[:255],
                supplier_siret=supplier_siret,
                purchase_amount=purchase_amount,
                purchase_category=purchase_category[:255] if purchase_category else None,
                buying_entity=buying_entity[:255] if buying_entity else None,
                purchase_year=year,
                siae=siae,
                company=company,
            )
        )

    wb.close()

    if purchases:
        from itertools import batched

        BATCH_SIZE = 1_000
        for batch in batched(purchases, BATCH_SIZE):
            Purchase.objects.bulk_create(list(batch))

    return errors, len(purchases), matched_rows


def generate_purchases_excel_template() -> bytes:
    """Generates a downloadable Excel template for purchase data import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dépenses achats"

    headers = [
        "Raison sociale du Fournisseur",
        "SIRET",
        "Dépense achat (€)",
        "Catégorie d'achat (optionnelle)",
        "Entité acheteuse (optionnelle)",
    ]

    # Header row with bold styling
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Example rows
    ws.append(["Entreprise Exemple SAS", "12345678901234", 15000, "Travaux d'entretien", "Direction des achats"])
    ws.append(["Prestataire Inclusif ESAT", "98765432109876", 8500, "Services informatiques", ""])

    # Column widths
    column_widths = [35, 18, 18, 30, 30]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Colonne", "Obligatoire", "Description"])
    ws2.append(["Raison sociale du Fournisseur", "Oui", "Nom de l'entreprise fournisseur"])
    ws2.append(["SIRET", "Oui", "14 chiffres sans espaces"])
    ws2.append(["Dépense achat (€)", "Oui", "Montant en euros (nombre positif)"])
    ws2.append(["Catégorie d'achat (optionnelle)", "Non", "Catégorie libre (ex : Travaux, SI, Nettoyage…)"])
    ws2.append(["Entité acheteuse (optionnelle)", "Non", "Service ou direction acheteuse"])

    import io

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
