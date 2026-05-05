import csv
import io
import json
import logging
from urllib.parse import urlencode

import openpyxl
import xlwt
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.text import slugify
from django.views import View
from django.views.generic import DetailView, FormView, UpdateView
from django_filters.views import FilterView

from content_manager.models import ContentPage, Tag
from lemarche.api.inclusive_potential.constants import PRESTA_MODE_DEFAULT, PRESTA_MODE_TO_SIAE_KINDS, RECOMMENDATIONS
from lemarche.api.inclusive_potential.utils import get_inclusive_potential_data
from lemarche.cms.models import ArticleList
from lemarche.perimeters.models import Perimeter
from lemarche.purchases.models import Purchase, get_sector_group_chart_data
from lemarche.sectors.models import Sector
from lemarche.siaes.constants import KIND_HANDICAP_LIST, KIND_INSERTION_LIST, LEGAL_FORM_CHOICES
from lemarche.siaes.models import Siae
from lemarche.tenders.models import Tender
from lemarche.users.models import User
from lemarche.utils.slug_matching import (
    RESOLUTION_STATUS_AMBIGUOUS,
    RESOLUTION_STATUS_ERROR,
    RESOLUTION_STATUS_RESOLVED,
    record_user_choices,
    resolve_column_header,
    resolve_perimeter,
    resolve_sector,
    resolve_sector_from_title,
)
from lemarche.www.dashboard.filters import PurchaseFilterSet
from lemarche.www.dashboard.forms import DisabledEmailEditForm, ProfileEditForm, PurchaseProjectFormSet


logger = logging.getLogger(__name__)


SLUG_RESSOURCES_CAT_SIAES = "solutions"
SLUG_RESSOURCES_CAT_BUYERS = "acheteurs"


class DashboardHomeView(LoginRequiredMixin, DetailView):
    context_object_name = "user"

    def get_object(self):
        return self.request.user

    def get_template_names(self):
        if self.request.user.kind == User.KIND_SIAE:
            return ["dashboard/home_siae.html"]
        return ["dashboard/home_buyer.html"]

    def get(self, request, *args, **kwargs):
        """
        Update 'dashboard_last_seen_date'
        """
        user = self.request.user
        if user.is_authenticated:
            User.objects.filter(id=user.id).update(dashboard_last_seen_date=timezone.now())
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # filter ressources by user kind
        category_slug = None
        if user.kind == User.KIND_SIAE:
            category_slug = SLUG_RESSOURCES_CAT_SIAES
        elif user.kind == User.KIND_BUYER:
            category_slug = SLUG_RESSOURCES_CAT_BUYERS

        # Get ContentPage under the ArticleList that has the slug "ressources"
        try:
            ressource_page = ArticleList.objects.get(slug="ressources")
            ressource_list = (
                ContentPage.objects.descendant_of(ressource_page)
                .live()
                .prefetch_related("tags")
                .order_by("-last_published_at")
            )
        except ArticleList.DoesNotExist:
            ressource_list = ContentPage.objects.none()

        if category_slug:
            try:
                tag = Tag.objects.get(slug=category_slug)
                ressource_list = ressource_list.filter(tags__in=[tag])
            except Exception:
                pass

        # set context ressources
        context["last_3_ressources"] = ressource_list[:3]

        # for specific users
        if user.kind == User.KIND_SIAE:
            siaes = user.siaes.all()
            if siaes:
                context["last_3_tenders"] = Tender.objects.filter_with_siaes(siaes).order_by_last_published()[:3]
        else:
            context["last_3_tenders"] = Tender.objects.filter(author=user).order_by_last_published()[:3]
            context["user_buyer_count"] = User.objects.filter(kind=User.KIND_BUYER).count()
            context["siae_count"] = Siae.objects.is_live().count()
            context["tender_count"] = Tender.objects.sent().count() + 30  # historic number (before form)
        return context


class InclusivePurchaseStatsDashboardView(LoginRequiredMixin, FilterView):
    filterset_class = PurchaseFilterSet
    template_name = "dashboard/inclusive_purchase_stats.html"

    def get_queryset(self):
        return Purchase.objects.get_purchase_for_user(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        if user.kind != User.KIND_BUYER and user.company is None:
            return context

        # Used to determine if the user has purchases, even if the filters leads to results
        context["unfiltered_qs_count"] = self.get_queryset().count()

        # get purchase stats for the user
        purchases_stats = self.filterset.qs.with_stats()
        total_purchases = purchases_stats["total_amount_annotated"]
        if total_purchases > 0:
            chart_data_inclusive = {
                "labels": ["Achats inclusifs", "Achats non inclusifs"],
                "dataset": [
                    purchases_stats["total_inclusive_amount_annotated"],
                    purchases_stats["total_amount_annotated"] - (purchases_stats["total_inclusive_amount_annotated"]),
                ],
            }
            chart_data_insertion_handicap = {
                "labels": ["Structures d'insertion (IAE)", "Structures du Handicap (STPA)"],
                "dataset": [
                    purchases_stats["total_insertion_amount_annotated"],
                    purchases_stats["total_handicap_amount_annotated"],
                ],
            }
            chart_data_siae_type = {
                "labels": [
                    kind
                    for kind in KIND_INSERTION_LIST + KIND_HANDICAP_LIST
                    if purchases_stats[f"total_purchases_by_kind_{kind}"] > 0
                ],
                "dataset": [
                    purchases_stats[f"total_purchases_by_kind_{kind}"]
                    for kind in KIND_INSERTION_LIST + KIND_HANDICAP_LIST
                    if purchases_stats[f"total_purchases_by_kind_{kind}"] > 0
                ],
            }
            purchase_categories = list(
                self.filterset.qs.filter(siae__isnull=False)
                .values_list("purchase_category", flat=True)
                .order_by("purchase_category")
                .distinct()
            )
            # Build list of (category, value) for non-zero values, sort desc and take top 40
            purchase_category_slugs = [
                slugify(purchase_category) or "none" for purchase_category in purchase_categories
            ]
            categories_with_values = []
            for purchase_category_slug in purchase_category_slugs:
                categories_with_values.append(
                    (purchase_category_slug, purchases_stats[f"total_purchases_by_category_{purchase_category_slug}"])
                )
            top_categories = sorted(categories_with_values, key=lambda item: item[1], reverse=True)[:40]
            chart_data_purchases_by_category = {
                "labels": [label if label != "none" else "Non renseigné" for label, _ in top_categories],
                "dataset": [value for _, value in top_categories],
            }

            buying_entities = list(
                self.filterset.qs.filter(siae__isnull=False)
                .values_list("buying_entity", flat=True)
                .order_by("buying_entity")
                .distinct()
            )
            buying_entity_slugs = [slugify(buying_entity) or "none" for buying_entity in buying_entities]
            chart_data_purchases_by_buying_entity = {
                "labels": [buying_entity or "Non renseigné" for buying_entity in buying_entities],
                "dataset": [
                    purchases_stats[f"total_purchases_by_buying_entity_{buying_entity_slug}"]
                    for buying_entity_slug in buying_entity_slugs
                    if purchases_stats[f"total_purchases_by_buying_entity_{buying_entity_slug}"] > 0
                ],
            }

            # QPV / ZRR breakdown: QPV only, ZRR only, both, neither
            qpv_zrr_labels = [
                "QPV uniquement",
                "ZRR uniquement",
                "QPV et ZRR",
                "Ni QPV ni ZRR",
            ]
            qpv_zrr_keys = [
                "total_purchases_qpv_only",
                "total_purchases_zrr_only",
                "total_purchases_qpv_and_zrr",
                "total_purchases_no_qpv_no_zrr",
            ]
            chart_data_qpv_zrr = {
                "labels": [label for label, key in zip(qpv_zrr_labels, qpv_zrr_keys) if purchases_stats[key] > 0],
                "dataset": [purchases_stats[key] for key in qpv_zrr_keys if purchases_stats[key] > 0],
            }

            # --- Taille de structure ---
            SIZE_ORDER = ["TPE (< 10 salariés)", "PME (≥ 10 salariés)", "Non renseigné"]
            size_rows = {row["size_category"]: row for row in self.filterset.qs.with_size_stats()}
            chart_data_size = {
                "labels": SIZE_ORDER,
                "amounts": [size_rows.get(cat, {}).get("total_amount", 0) for cat in SIZE_ORDER],
                "supplier_counts": [size_rows.get(cat, {}).get("supplier_count", 0) for cat in SIZE_ORDER],
            }

            # --- Statut juridique ---
            legal_form_label_map = dict(LEGAL_FORM_CHOICES)
            legal_form_rows = list(self.filterset.qs.with_legal_form_stats())
            TOP_N = 10
            top_rows = legal_form_rows[:TOP_N]
            other_rows = legal_form_rows[TOP_N:]
            lf_labels = [
                legal_form_label_map.get(r["siae__legal_form"], r["siae__legal_form"] or "Non renseigné")
                for r in top_rows
            ]
            lf_amounts = [r["total_amount"] for r in top_rows]
            lf_supplier_counts = [r["supplier_count"] for r in top_rows]
            if other_rows:
                lf_labels.append("Autres")
                lf_amounts.append(sum(r["total_amount"] for r in other_rows))
                lf_supplier_counts.append(sum(r["supplier_count"] for r in other_rows))
            chart_data_legal_form = {
                "labels": lf_labels,
                "amounts": lf_amounts,
                "supplier_counts": lf_supplier_counts,
            }

            # --- Secteurs d'activité ---
            chart_data_sector_group = get_sector_group_chart_data(self.filterset.qs)

            # --- Cartographie régionale ---
            region_rows = list(self.filterset.qs.with_region_stats())
            total_inclusive = purchases_stats["total_inclusive_amount_annotated"]
            region_table_rows = []
            for r in region_rows:
                label = r["siae__region"] or "Non renseigné"
                amount = r["total_amount"]
                pct = round(amount * 100 / total_inclusive, 1) if total_inclusive else 0
                region_table_rows.append((label, amount, pct, r["supplier_count"]))
            chart_data_region = {
                "labels": [row[0] for row in region_table_rows],
                "amounts": [row[1] for row in region_table_rows],
                "supplier_counts": [row[3] for row in region_table_rows],
                "rows": region_table_rows,
            }

            context.update(
                {
                    "total_purchases": purchases_stats["total_amount_annotated"],
                    "total_suppliers": purchases_stats["total_suppliers_annotated"],
                    "total_inclusive_suppliers": purchases_stats["total_inclusive_suppliers_annotated"],
                    "total_inclusive_purchases": purchases_stats["total_inclusive_amount_annotated"],
                    "total_insertion_purchases": purchases_stats["total_insertion_amount_annotated"],
                    "total_handicap_purchases": purchases_stats["total_handicap_amount_annotated"],
                    "total_inclusive_purchases_percentage": round(
                        purchases_stats["total_inclusive_amount_annotated"] * 100 / total_purchases,
                        2,
                    ),
                    "total_insertion_purchases_percentage": round(
                        purchases_stats["total_insertion_amount_annotated"] * 100 / total_purchases,
                        2,
                    ),
                    "total_handicap_purchases_percentage": round(
                        purchases_stats["total_handicap_amount_annotated"] * 100 / total_purchases,
                        2,
                    ),
                    "chart_data_inclusive": chart_data_inclusive,
                    "chart_data_insertion_handicap": chart_data_insertion_handicap,
                    "chart_data_siae_type": chart_data_siae_type,
                    "chart_data_purchases_by_category": chart_data_purchases_by_category,
                    "chart_data_purchases_by_buying_entity": chart_data_purchases_by_buying_entity,
                    "chart_data_qpv_zrr": chart_data_qpv_zrr,
                    "chart_data_size": chart_data_size,
                    "chart_data_legal_form": chart_data_legal_form,
                    "chart_data_region": chart_data_region,
                    "chart_data_sector_group": chart_data_sector_group,
                }
            )
        return context


class InclusivePurchaseExportView(LoginRequiredMixin, View):
    """
    Export filtered purchase data as CSV or Excel (.xls).
    Accepts the same query params as InclusivePurchaseStatsDashboardView.
    Usage: ?format=csv  or  ?format=excel
    """

    LEGAL_FORM_LABELS = dict(LEGAL_FORM_CHOICES)

    EXPORT_HEADERS = [
        "Fournisseur",
        "SIRET",
        "Montant (€)",
        "Année",
        "Catégorie d'achat",
        "Entité acheteuse",
        "Structure inclusive",
        "Type de structure",
        "Statut juridique",
        "Région",
        "Taille",
        "QPV",
        "ZRR",
    ]

    def _get_size_label(self, siae):
        if siae is None:
            return ""
        insertion = siae.employees_insertion_count
        if insertion is None and siae.c2_etp_count is not None:
            insertion = round(siae.c2_etp_count)
        permanent = siae.employees_permanent_count or 0
        if insertion is None and siae.employees_permanent_count is None:
            return "Non renseigné"
        total = (insertion or 0) + permanent
        return "TPE (< 10 salariés)" if total < 10 else "PME (≥ 10 salariés)"

    def _build_rows(self, purchases):
        rows = []
        for purchase in purchases.select_related("siae"):
            siae = purchase.siae
            rows.append(
                [
                    purchase.supplier_name,
                    purchase.supplier_siret,
                    float(purchase.purchase_amount),
                    purchase.purchase_year,
                    purchase.purchase_category or "",
                    purchase.buying_entity or "",
                    siae.name if siae else "",
                    siae.kind if siae else "",
                    self.LEGAL_FORM_LABELS.get(siae.legal_form, siae.legal_form or "") if siae else "",
                    siae.region if siae else "",
                    self._get_size_label(siae),
                    "Oui" if siae and siae.is_qpv else ("Non" if siae else ""),
                    "Oui" if siae and siae.is_zrr else ("Non" if siae else ""),
                ]
            )
        return rows

    def get(self, request):
        filterset = PurchaseFilterSet(
            data=request.GET,
            queryset=Purchase.objects.get_purchase_for_user(request.user),
        )
        purchases = filterset.qs
        export_format = request.GET.get("format", "csv")

        if export_format == "excel":
            return self._export_excel(purchases)
        return self._export_csv(purchases)

    def _export_csv(self, purchases):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="achats_inclusifs.csv"'
        response.write("\ufeff")  # BOM for Excel UTF-8 compatibility
        writer = csv.writer(response, delimiter=";")
        writer.writerow(self.EXPORT_HEADERS)
        for row in self._build_rows(purchases):
            writer.writerow(row)
        return response

    def _export_excel(self, purchases):
        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("Achats inclusifs")

        bold = xlwt.XFStyle()
        bold.font.bold = True
        normal = xlwt.XFStyle()
        normal.alignment.wrap = 1

        for col, header in enumerate(self.EXPORT_HEADERS):
            ws.write(0, col, header, bold)

        for row_idx, row in enumerate(self._build_rows(purchases), start=1):
            for col_idx, value in enumerate(row):
                ws.write(row_idx, col_idx, value, normal)

        response = HttpResponse(content_type="application/vnd.ms-excel")
        response["Content-Disposition"] = 'attachment; filename="achats_inclusifs.xls"'
        wb.save(response)
        return response


class ProfileEditView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    form_class = ProfileEditForm
    template_name = "dashboard/profile_edit.html"
    success_message = "Votre profil a été mis à jour."
    success_url = reverse_lazy("dashboard:home")

    def get_object(self):
        return self.request.user


class DisabledEmailEditView(LoginRequiredMixin, SuccessMessageMixin, FormView):
    form_class = DisabledEmailEditForm
    template_name = "dashboard/disabled_email_edit.html"
    success_url = reverse_lazy("dashboard:notifications_edit")
    success_message = "Vos préférences de notifications ont été mises à jour."

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.save()
        return super().form_valid(form)


def _build_search_urls(sector: Sector, perimeter: Perimeter | None, presta_mode: str | None = None) -> dict:
    """Build search URLs for each indicator, pointing to the Marché search page."""
    base_params = [("sectors", sector.slug)]
    if perimeter:
        base_params.append(("perimeters", perimeter.slug))

    presta_kinds = PRESTA_MODE_TO_SIAE_KINDS.get(presta_mode) if presta_mode else None

    def kind_params(base_kinds=None):
        """Return kind URL params, intersected with active presta_mode if any."""
        if base_kinds is None:
            kinds = presta_kinds or []
        elif presta_kinds is not None:
            kinds = [k for k in base_kinds if k in presta_kinds]
        else:
            kinds = base_kinds
        return [("kind", k) for k in kinds]

    def url(kind_filter=None, extra_params=None):
        params = base_params + kind_params(kind_filter) + (extra_params or [])
        return f"/prestataires/?{urlencode(params)}#searchResults"

    return {
        "all": url(),
        "insertion": url(KIND_INSERTION_LIST),
        "handicap": url(KIND_HANDICAP_LIST),
        "local": url(extra_params=[("local", "True")]),
        "super_badge": url(extra_params=[("super_badge", "True")]),
        "won_contract": url(extra_params=[("has_won_contract", "True")]),
    }


def _analyze_purchase_project(
    titre: str, sector: Sector, perimeter: Perimeter | None, budget: int | None, presta_mode: str = PRESTA_MODE_DEFAULT
) -> dict:
    """Run the inclusive potential analysis for a single purchase project."""
    try:
        potential_data, analysis_data = get_inclusive_potential_data(sector, perimeter, budget, presta_mode)
    except Exception:
        logger.exception("Erreur lors de l'analyse du potentiel inclusif pour '%s'", titre)
        return {
            "titre": titre,
            "error": "Une erreur technique s'est produite lors de l'analyse. Veuillez réessayer.",
        }

    search_urls = _build_search_urls(sector, perimeter, presta_mode)
    result = {
        "titre": titre,
        "secteur_name": sector.name,
        "perimeter_name": perimeter.name if perimeter else "Toute la France",
        "montant": f"{budget:,}".replace(",", " ") if budget else None,
        "potential_siaes": potential_data.potential_siaes,
        "insertion_siaes": potential_data.insertion_siaes,
        "handicap_siaes": potential_data.handicap_siaes,
        "local_siaes": potential_data.local_siaes,
        "siaes_with_super_badge": potential_data.siaes_with_super_badge,
        "siaes_with_won_contract": potential_data.siaes_with_won_contract,
        "employees_insertion_average": potential_data.employees_insertion_average,
        "employees_permanent_average": potential_data.employees_permanent_average,
        "ca_average": None,
        "eco_dependency": None,
        "recommendation_title": None,
        "recommendation_response": None,
        "search_urls": search_urls,
        "error": None,
    }

    if budget and "recommendation" in analysis_data:
        result["ca_average"] = analysis_data.get("ca_average")
        result["eco_dependency"] = analysis_data.get("eco_dependency")
        recommendation = analysis_data["recommendation"]
        result["recommendation_title"] = recommendation.get("title")
        result["recommendation_response"] = recommendation.get("response")
        result["recommendation_key"] = _RECOMMENDATION_TITLE_TO_KEY.get(result["recommendation_title"])

    return result


_RECOMMENDATION_TITLE_TO_KEY = {rec["title"]: key for key, rec in RECOMMENDATIONS.items()}

INCLUSIVE_POTENTIAL_ANALYSIS_TEMPLATE = "dashboard/inclusive_potential_analysis.html"


def _run_excel_analysis(
    request, resolved_projects: list, unresolvable_count: int, presta_mode: str = PRESTA_MODE_DEFAULT
):
    """Lance l'analyse IPA sur les projets résolus et retourne la réponse rendue."""
    if presta_mode not in PRESTA_MODE_TO_SIAE_KINDS:
        presta_mode = PRESTA_MODE_DEFAULT

    results = []
    raw_projects_session = []

    for project in resolved_projects:
        sector_slug = project["sector_slug"]
        perimeter_result = project["perimeter_result"]
        montant = project["montant"]
        titre = project["titre"]

        try:
            sector = Sector.objects.get(slug=sector_slug)
        except Sector.DoesNotExist:
            continue

        perimeter = None
        if not perimeter_result.get("france_entiere") and perimeter_result.get("slug"):
            try:
                perimeter = Perimeter.objects.get(slug=perimeter_result["slug"])
            except Perimeter.DoesNotExist:
                continue

        result = _analyze_purchase_project(titre, sector, perimeter, montant, presta_mode)
        results.append(result)
        raw_projects_session.append(
            {
                "titre": titre,
                "montant": montant,
                "secteur_slug": sector_slug,
                "perimeter_slug": perimeter_result.get("slug"),
                "france_entiere": perimeter_result.get("france_entiere", False),
                "input_mode": "excel",
            }
        )

    context_extra = {}
    if unresolvable_count:
        context_extra["import_errors"] = [f"⚠️ {unresolvable_count} projet(s) ignoré(s) : correspondance introuvable."]

    if not results:
        error_msg = "Aucun projet analysable trouvé."
        if unresolvable_count:
            error_msg += f" {unresolvable_count} projet(s) ignoré(s) : correspondance introuvable."
        return render(
            request,
            INCLUSIVE_POTENTIAL_ANALYSIS_TEMPLATE,
            {
                "sectors": list(Sector.objects.form_filter_queryset()),
                "formset": PurchaseProjectFormSet(prefix="form"),
                "import_errors": [error_msg],
                "mode": "excel",
            },
        )

    for i, result in enumerate(results):
        if not result.get("error"):
            result["detail_url"] = reverse("dashboard:inclusive_potential_project_detail", args=[i])

    by_sector, by_perimeter = _aggregate_results(results)
    request.session["ipa_excel_results"] = [{k: v for k, v in r.items() if k != "search_urls"} for r in results]
    request.session["ipa_raw_projects"] = raw_projects_session
    request.session["ipa_presta_mode"] = presta_mode

    return render(
        request,
        INCLUSIVE_POTENTIAL_ANALYSIS_TEMPLATE,
        {
            "sectors": list(Sector.objects.form_filter_queryset()),
            "formset": PurchaseProjectFormSet(prefix="form"),
            "results": results,
            "results_by_sector": by_sector,
            "results_by_perimeter": by_perimeter,
            "mode": "excel",
            "presta_mode": presta_mode,
            "presta_mode_choices": list(PRESTA_MODE_TO_SIAE_KINDS.keys()),
            **context_extra,
        },
    )


def _parse_excel_projects(file) -> list[dict]:
    """Parse an uploaded Excel file into a list of raw project dicts.
    Applies fuzzy matching on column headers to tolerate non-standard names.
    Raises ValueError with a user-friendly message on validation failure.
    """

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception:
        raise ValueError("Le fichier n'est pas un fichier Excel valide (.xlsx).")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Le fichier est vide.")

    # Recherche automatique de la ligne d'en-têtes dans les 10 premières lignes
    # (certains fichiers ont des lignes de légende ou d'instructions avant les colonnes)
    header_row_idx = None
    col = {}
    for idx, row in enumerate(rows[:10]):
        candidate = {}
        for i, h in enumerate(row):
            if not h:
                continue
            resolved = resolve_column_header(str(h))
            if resolved and resolved not in candidate:
                candidate[resolved] = i
        if "titre" in candidate:
            header_row_idx = idx
            col = candidate
            break

    if header_row_idx is None:
        raise ValueError(
            "Impossible de trouver la ligne d'en-têtes. "
            "Assurez-vous que votre fichier contient une colonne 'Titre' (ou équivalent)."
        )

    data_rows = [r for r in rows[header_row_idx + 1 :] if any(r)]

    if not data_rows:
        raise ValueError("Le fichier ne contient aucune ligne de données.")

    projects = []
    for row_num, row in enumerate(data_rows, start=2):

        def cell(name, _row=row):
            idx = col.get(name)
            return str(_row[idx]).strip() if idx is not None and _row[idx] is not None else ""

        montant = None
        if col.get("montant") is not None and row[col["montant"]] is not None:
            try:
                montant = int(float(str(row[col["montant"]])))
            except (ValueError, TypeError):
                montant = None

        projects.append(
            {
                "row": row_num,
                "titre": cell("titre"),
                "description": cell("description"),
                "secteur_raw": cell("secteur"),
                "perimetre_raw": cell("perimetre_geographique"),
                "montant": montant,
            }
        )

    return projects


def _aggregate_results(results: list[dict]) -> tuple[dict, dict]:
    """Aggregate valid results by sector name and by perimeter name.
    Returns (by_sector, by_perimeter) — each value is a dict with projects + stats.
    """
    by_sector: dict = {}
    by_perimeter: dict = {}

    for result in results:
        if result.get("error"):
            continue

        for key, group in [
            (result["secteur_name"], by_sector),
            (result["perimeter_name"], by_perimeter),
        ]:
            if key not in group:
                group[key] = {"projects": [], "total_structures": 0, "total_montant": 0, "project_count": 0}

            entry = group[key]
            entry["projects"].append(result)
            entry["total_structures"] += result["potential_siaes"]
            entry["project_count"] += 1
            if result.get("montant"):
                try:
                    entry["total_montant"] += int(result["montant"].replace(" ", "").replace(" ", ""))
                except (ValueError, AttributeError):
                    pass

    sort_key = lambda x: x[1]["total_structures"]  # noqa: E731
    by_sector = dict(sorted(by_sector.items(), key=sort_key, reverse=True))
    by_perimeter = dict(sorted(by_perimeter.items(), key=sort_key, reverse=True))

    return by_sector, by_perimeter


class InclusivePotentialAnalysisView(LoginRequiredMixin, View):
    template_name = "dashboard/inclusive_potential_analysis.html"

    def _get_context(
        self,
        formset=None,
        results=None,
        results_by_sector=None,
        results_by_perimeter=None,
        import_errors=None,
        mode="manual",
        presta_mode=PRESTA_MODE_DEFAULT,
    ):
        return {
            "sectors": list(Sector.objects.form_filter_queryset()),
            "formset": formset or PurchaseProjectFormSet(prefix="form"),
            "results": results,
            "results_by_sector": results_by_sector,
            "results_by_perimeter": results_by_perimeter,
            "import_errors": import_errors,
            "mode": mode,
            "presta_mode": presta_mode,
            "presta_mode_choices": list(PRESTA_MODE_TO_SIAE_KINDS.keys()),
        }

    def get(self, request):
        if request.GET.get("from_detail"):
            response = self._restore_excel_results(request)
            if response:
                return response
        return render(request, self.template_name, self._get_context())

    def _restore_excel_results(self, request):
        """Restaure les résultats Excel depuis la session (retour depuis la page détail projet)."""
        raw_projects = request.session.get("ipa_raw_projects", [])
        if not raw_projects or raw_projects[0].get("input_mode") != "excel":
            return None

        presta_mode = request.session.get("ipa_presta_mode", PRESTA_MODE_DEFAULT)
        if presta_mode not in PRESTA_MODE_TO_SIAE_KINDS:
            presta_mode = PRESTA_MODE_DEFAULT

        results = []
        for project in raw_projects:
            try:
                sector = Sector.objects.get(slug=project["secteur_slug"])
            except Sector.DoesNotExist:
                continue
            perimeter = None
            if project.get("perimeter_slug") and not project.get("france_entiere"):
                try:
                    perimeter = Perimeter.objects.get(slug=project["perimeter_slug"])
                except Perimeter.DoesNotExist:
                    continue
            result = _analyze_purchase_project(
                project["titre"], sector, perimeter, project.get("montant"), presta_mode
            )
            results.append(result)

        if not results:
            return None

        for i, result in enumerate(results):
            if not result.get("error"):
                result["detail_url"] = reverse("dashboard:inclusive_potential_project_detail", args=[i])

        by_sector, by_perimeter = _aggregate_results(results)
        return render(
            request,
            self.template_name,
            self._get_context(
                results=results,
                results_by_sector=by_sector,
                results_by_perimeter=by_perimeter,
                mode="excel",
                presta_mode=presta_mode,
            ),
        )

    def post(self, request):
        mode = request.POST.get("mode", "manual")
        if mode == "reanalyze":
            return self._handle_reanalyze(request)
        if mode == "excel":
            return self._handle_excel_import(request)
        return self._handle_manual_form(request)

    def _handle_manual_form(self, request):
        formset = PurchaseProjectFormSet(request.POST, prefix="form")
        if not formset.is_valid():
            return render(
                request,
                self.template_name,
                self._get_context(formset=formset, mode="manual"),
            )

        presta_mode = request.POST.get("presta_mode", PRESTA_MODE_DEFAULT)
        if presta_mode not in PRESTA_MODE_TO_SIAE_KINDS:
            presta_mode = PRESTA_MODE_DEFAULT

        raw_projects = []
        results = []
        for form in formset:
            if not form.cleaned_data:
                continue
            titre = form.cleaned_data["titre"]
            sector = form.cleaned_data["secteur"]
            montant = form.cleaned_data.get("montant")
            perimeter_slug = form.cleaned_data.get("perimeter_slug")
            france_entiere = form.cleaned_data.get("france_entiere", False)

            perimeter = None
            if perimeter_slug and not france_entiere:
                try:
                    perimeter = Perimeter.objects.get(slug=perimeter_slug)
                except Perimeter.DoesNotExist:
                    results.append({"titre": titre, "error": f"Périmètre '{perimeter_slug}' introuvable."})
                    continue

            raw_projects.append(
                {
                    "titre": titre,
                    "secteur_slug": sector.slug,
                    "perimeter_slug": perimeter.slug if perimeter else None,
                    "france_entiere": france_entiere,
                    "montant": montant,
                    "input_mode": "manual",
                }
            )
            result = _analyze_purchase_project(titre, sector, perimeter, montant, presta_mode)
            results.append(result)

        request.session["ipa_raw_projects"] = raw_projects

        return render(
            request,
            self.template_name,
            self._get_context(formset=formset, results=results, mode="manual", presta_mode=presta_mode),
        )

    def _handle_reanalyze(self, request):
        """Re-run the analysis from session data with a new presta_mode."""
        raw_projects = request.session.get("ipa_raw_projects")
        if not raw_projects:
            return HttpResponseRedirect(reverse("dashboard:inclusive_potential_analysis"))

        presta_mode = request.POST.get("presta_mode", PRESTA_MODE_DEFAULT)
        if presta_mode not in PRESTA_MODE_TO_SIAE_KINDS:
            presta_mode = PRESTA_MODE_DEFAULT

        input_mode = raw_projects[0].get("input_mode", "manual") if raw_projects else "manual"

        results = []
        for project in raw_projects:
            titre = project["titre"]
            montant = project["montant"]
            france_entiere = project.get("france_entiere", False)

            try:
                sector = Sector.objects.get(slug=project["secteur_slug"])
            except Sector.DoesNotExist:
                results.append({"titre": titre, "error": f"Secteur '{project['secteur_slug']}' introuvable."})
                continue

            perimeter = None
            if project.get("perimeter_slug") and not france_entiere:
                try:
                    perimeter = Perimeter.objects.get(slug=project["perimeter_slug"])
                except Perimeter.DoesNotExist:
                    results.append({"titre": titre, "error": f"Périmètre '{project['perimeter_slug']}' introuvable."})
                    continue

            result = _analyze_purchase_project(titre, sector, perimeter, montant, presta_mode)
            results.append(result)

        request.session["ipa_raw_projects"] = raw_projects

        if input_mode == "excel":
            for i, result in enumerate(results):
                if not result.get("error"):
                    result["detail_url"] = reverse("dashboard:inclusive_potential_project_detail", args=[i])
            request.session["ipa_excel_results"] = [
                {k: v for k, v in r.items() if k != "search_urls"} for r in results
            ]
            request.session["ipa_presta_mode"] = presta_mode
            by_sector, by_perimeter = _aggregate_results(results)
            return render(
                request,
                self.template_name,
                self._get_context(
                    results=results,
                    results_by_sector=by_sector,
                    results_by_perimeter=by_perimeter,
                    mode="excel",
                    presta_mode=presta_mode,
                ),
            )

        return render(
            request,
            self.template_name,
            self._get_context(results=results, mode="manual", presta_mode=presta_mode),
        )

    def _handle_excel_import(self, request):
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            return render(
                request,
                self.template_name,
                self._get_context(import_errors=["Aucun fichier sélectionné."], mode="excel"),
            )

        try:
            projects = _parse_excel_projects(excel_file)
        except ValueError as e:
            return render(
                request,
                self.template_name,
                self._get_context(import_errors=[str(e)], mode="excel"),
            )

        presta_mode = request.POST.get("presta_mode", PRESTA_MODE_DEFAULT)
        if presta_mode not in PRESTA_MODE_TO_SIAE_KINDS:
            presta_mode = PRESTA_MODE_DEFAULT

        resolved_projects = []
        ambiguous_items = []
        unresolvable_count = 0

        for project in projects:
            titre = project["titre"]
            row_num = project["row"]

            if not titre:
                unresolvable_count += 1
                continue

            sector_result = resolve_sector(project["secteur_raw"], user=request.user)

            # Si le secteur est introuvable, tenter l'inférence depuis le titre/description
            if sector_result["status"] == RESOLUTION_STATUS_ERROR and titre:
                sector_result = resolve_sector_from_title(titre, project.get("description", ""))

            # Si le périmètre est vide (colonne absente ou non renseignée), envoyer en validation manuelle
            if not project["perimetre_raw"]:
                perimeter_result = {
                    "status": RESOLUTION_STATUS_AMBIGUOUS,
                    "slug": None,
                    "france_entiere": False,
                    "candidates": [],
                    "source": "",
                }
            else:
                perimeter_result = resolve_perimeter(project["perimetre_raw"], user=request.user)

            sector_status = sector_result["status"]
            perimeter_status = perimeter_result["status"]

            if sector_status == RESOLUTION_STATUS_RESOLVED and perimeter_status == RESOLUTION_STATUS_RESOLVED:
                resolved_projects.append(
                    {**project, "sector_slug": sector_result["slug"], "perimeter_result": perimeter_result}
                )
            elif sector_status == RESOLUTION_STATUS_ERROR or perimeter_status == RESOLUTION_STATUS_ERROR:
                unresolvable_count += 1
            else:
                ambiguous_items.append(
                    {
                        "row": row_num,
                        "titre": titre,
                        "montant": project["montant"],
                        "description": project.get("description", ""),
                        "sector_raw": project["secteur_raw"],
                        "perimeter_raw": project["perimetre_raw"],
                        "sector_result": sector_result,
                        "perimeter_result": perimeter_result,
                    }
                )

        if ambiguous_items:
            request.session["ipa_pending_projects"] = {
                "resolved": resolved_projects,
                "ambiguous": ambiguous_items,
                "unresolvable_count": unresolvable_count,
                "presta_mode": presta_mode,
            }
            return HttpResponseRedirect(reverse("dashboard:slug_mapping_validation"))

        return _run_excel_analysis(request, resolved_projects, unresolvable_count, presta_mode)


class InclusivePotentialProjectDetailView(LoginRequiredMixin, View):
    """Affiche l'analyse détaillée d'un projet d'achat issu d'un import Excel IPA.

    Lit le projet à l'index donné dans ipa_raw_projects (session), relance l'analyse
    et affiche le résultat sur le même template que le mode saisie manuelle.
    """

    template_name = "dashboard/inclusive_potential_analysis.html"

    def get(self, request, index: int):
        raw_projects = request.session.get("ipa_raw_projects", [])

        if not raw_projects or index >= len(raw_projects):
            return HttpResponseRedirect(reverse("dashboard:inclusive_potential_analysis"))

        project = raw_projects[index]

        if project.get("input_mode") != "excel":
            return HttpResponseRedirect(reverse("dashboard:inclusive_potential_analysis"))

        presta_mode = request.session.get("ipa_presta_mode", PRESTA_MODE_DEFAULT)
        if presta_mode not in PRESTA_MODE_TO_SIAE_KINDS:
            presta_mode = PRESTA_MODE_DEFAULT

        try:
            sector = Sector.objects.get(slug=project["secteur_slug"])
        except Sector.DoesNotExist:
            return HttpResponseRedirect(reverse("dashboard:inclusive_potential_analysis"))

        perimeter = None
        if project.get("perimeter_slug") and not project.get("france_entiere"):
            try:
                perimeter = Perimeter.objects.get(slug=project["perimeter_slug"])
            except Perimeter.DoesNotExist:
                return HttpResponseRedirect(reverse("dashboard:inclusive_potential_analysis"))

        result = _analyze_purchase_project(project["titre"], sector, perimeter, project.get("montant"), presta_mode)

        return render(
            request,
            self.template_name,
            {
                "results": [result],
                "mode": "manual",
                "presta_mode": presta_mode,
                "presta_mode_choices": list(PRESTA_MODE_TO_SIAE_KINDS.keys()),
                "is_excel_project_detail": True,
                "back_url": reverse("dashboard:inclusive_potential_analysis") + "?from_detail=1",
                "sectors": [],
                "formset": None,
            },
        )


class SlugMappingValidationView(LoginRequiredMixin, View):
    """Affiche la page de validation des correspondances ambiguës issues de l'import Excel IPA.

    GET  — affiche les items ambigus stockés en session pour validation manuelle.
    POST — collecte les choix utilisateur, enregistre en cache, lance l'analyse.
    """

    template_name = "dashboard/inclusive_potential_mapping_validation.html"

    def get(self, request, *args, **kwargs):
        pending = request.session.get("ipa_pending_projects")
        if not pending:
            return HttpResponseRedirect(reverse("dashboard:inclusive_potential_analysis"))
        sectors_json = json.dumps(list(Sector.objects.form_filter_queryset().values("slug", "name")))
        return render(
            request,
            self.template_name,
            {
                "ambiguous_items": pending["ambiguous"],
                "ambiguous_items_json": json.dumps(pending["ambiguous"]),
                "unresolvable_count": pending.get("unresolvable_count", 0),
                "sectors_json": sectors_json,
            },
        )

    def post(self, request, *args, **kwargs):
        pending = request.session.get("ipa_pending_projects")
        if not pending:
            return HttpResponseRedirect(reverse("dashboard:inclusive_potential_analysis"))

        ambiguous_items = pending["ambiguous"]
        resolved_projects = list(pending["resolved"])
        unresolvable_count = pending.get("unresolvable_count", 0)
        presta_mode = pending.get("presta_mode", PRESTA_MODE_DEFAULT)

        # L'utilisateur veut analyser uniquement les projets déjà résolus, sans traiter les ambigus
        if request.POST.get("ignore_remaining") == "1":
            del request.session["ipa_pending_projects"]
            return _run_excel_analysis(
                request, resolved_projects, unresolvable_count + len(ambiguous_items), presta_mode
            )

        user_choices = []

        for item in ambiguous_items:
            row = str(item["row"])
            sector_slug = next((v.strip() for v in request.POST.getlist(f"sector_{row}") if v.strip()), "")
            perimeter_slug = next((v.strip() for v in request.POST.getlist(f"perimeter_{row}") if v.strip()), "")
            france_entiere = request.POST.get(f"france_entiere_{row}") == "1"

            if not sector_slug or (not perimeter_slug and not france_entiere):
                unresolvable_count += 1
                continue

            resolved_projects.append(
                {
                    "titre": item["titre"],
                    "montant": item["montant"],
                    "description": item.get("description", ""),
                    "sector_slug": sector_slug,
                    "perimeter_result": {"slug": perimeter_slug or None, "france_entiere": france_entiere},
                }
            )

            try:
                sector_confidence = float(request.POST.get(f"sector_confidence_{row}", 1.0))
            except (ValueError, TypeError):
                sector_confidence = 1.0
            try:
                perimeter_confidence = float(request.POST.get(f"perimeter_confidence_{row}", 1.0))
            except (ValueError, TypeError):
                perimeter_confidence = 1.0

            if item["sector_raw"] and sector_slug:
                user_choices.append(
                    {
                        "raw_value": item["sector_raw"],
                        "kind": "sector",
                        "resolved_slug": sector_slug,
                        "confidence": sector_confidence,
                    }
                )
            if item["perimeter_raw"] and perimeter_slug:
                user_choices.append(
                    {
                        "raw_value": item["perimeter_raw"],
                        "kind": "perimeter",
                        "resolved_slug": perimeter_slug,
                        "confidence": perimeter_confidence,
                    }
                )

        if user_choices:
            try:
                record_user_choices(user_choices, request.user)
            except Exception:
                # Ne jamais bloquer l'analyse si l'enregistrement du cache échoue
                logger.exception("Erreur lors de l'enregistrement des choix de matching en cache")

        del request.session["ipa_pending_projects"]
        return _run_excel_analysis(request, resolved_projects, unresolvable_count, presta_mode)


@login_required
def inclusive_potential_excel_template(request):
    """Generate and return a pre-formatted Excel template for batch import.

    Sheet 1 — Projets d'achat : 2 example rows using real sector slugs from the DB.
    Sheet 2 — Secteurs disponibles : full list of valid sector slugs + names.
    """

    wb = openpyxl.Workbook()

    # Feuille 1 : modèle à remplir
    ws = wb.active
    ws.title = "Projets d'achat"

    ws.append(["Titre du projet", "Catégorie achat", "Localisation", "Montant €", "Description"])

    example_sectors = list(Sector.objects.form_filter_queryset().values_list("slug", "name")[:2])
    if example_sectors:
        ws.append(
            [
                f"Prestations de {example_sectors[0][1][:40]}",
                example_sectors[0][1],
                "Paris",
                80000,
                "Description optionnelle",
            ]
        )
    if len(example_sectors) > 1:
        ws.append(
            [
                f"Prestations de {example_sectors[1][1][:40]}",
                example_sectors[1][1],
                "France entière",
                45000,
                "",
            ]
        )

    # Feuille 2 : référentiel des secteurs (nom + slug)
    ws2 = wb.create_sheet("Secteurs disponibles")
    ws2.append(["Nom du secteur (à saisir dans Catégorie achat)", "Slug technique (alternative)"])
    for slug, name in Sector.objects.form_filter_queryset().values_list("slug", "name"):
        ws2.append([name, slug])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="modele_analyse_potentiel_inclusif.xlsx"'
    return response


@login_required
def inclusive_potential_excel_export(request):
    """Export the last Excel analysis results as a downloadable .xlsx file."""
    results = request.session.get("ipa_excel_results")
    if not results:
        return HttpResponseRedirect(reverse("dashboard:inclusive_potential_analysis"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analyse potentiel inclusif"

    ws.append(
        [
            "Titre du projet d'achat",
            "Catégorie achat",
            "Périmètre",
            "Montant (€)",
            "Structures potentielles",
            "Structures d'insertion",
            "Structures du handicap",
            "Structures locales",
            "Super prestataires",
            "Marchés publics remportés",
            "ETP insertion (moy.)",
            "ETP permanents (moy.)",
            "Dépendance économique (%)",
            "CA moyen des structures (€)",
            "Recommandation",
        ]
    )

    for result in results:
        montant_raw = None
        if result.get("montant"):
            try:
                montant_raw = int(result["montant"].replace("\xa0", "").replace(" ", ""))
            except (ValueError, AttributeError):
                montant_raw = result["montant"]

        ws.append(
            [
                result.get("titre"),
                result.get("secteur_name"),
                result.get("perimeter_name"),
                montant_raw,
                result.get("potential_siaes"),
                result.get("insertion_siaes"),
                result.get("handicap_siaes"),
                result.get("local_siaes"),
                result.get("siaes_with_super_badge"),
                result.get("siaes_with_won_contract"),
                result.get("employees_insertion_average"),
                result.get("employees_permanent_average"),
                result.get("eco_dependency"),
                result.get("ca_average"),
                result.get("recommendation_title"),
            ]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="analyse_potentiel_inclusif.xlsx"'
    return response
