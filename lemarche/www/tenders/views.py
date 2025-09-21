import csv
import os
from datetime import timedelta

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Prefetch
from django.forms import formset_factory
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.views.generic import DetailView, ListView, UpdateView, View
from django.views.generic.edit import FormMixin, FormView
from formtools.wizard.views import SessionWizardView

from lemarche.siaes.models import Siae
from lemarche.siaes.tasks import send_reminder_email_to_siae
from lemarche.tenders import constants as tender_constants
from lemarche.tenders.enums import TenderSourcesChoices
from lemarche.tenders.forms import QuestionAnswerForm, SiaeSelectionForm
from lemarche.tenders.models import (
    QuestionAnswer,
    SuggestedQuestion,
    Tender,
    TenderInstruction,
    TenderQuestion,
    TenderSiae,
    TenderStepsData,
)
from lemarche.users import constants as user_constants
from lemarche.users.models import User
from lemarche.utils import constants, settings_context_processors
from lemarche.utils.data import get_choice
from lemarche.utils.emails import add_to_contact_list
from lemarche.utils.export import generate_siae_row
from lemarche.utils.mixins import (
    SesameSiaeMemberRequiredMixin,
    SesameTenderAuthorRequiredMixin,
    SiaeUserRequiredOrTenderSiaeUUIDParamMixin,
    TenderAuthorOrAdminRequiredIfNotSentMixin,
    TenderAuthorOrAdminRequiredMixin,
)
from lemarche.utils.urls import get_domain_url
from lemarche.www.siaes.forms import SiaeFilterForm
from lemarche.www.tenders.forms import (
    SiaeSelectFieldsForm,
    TenderCreateStepConfirmationForm,
    TenderCreateStepContactForm,
    TenderCreateStepDetailForm,
    TenderCreateStepGeneralForm,
    TenderCreateStepSignInForm,
    TenderCreateStepSurveyForm,
    TenderDetailGetParams,
    TenderFilterForm,
    TenderReminderForm,
    TenderSiaeSurveyTransactionedForm,
    TenderSurveyTransactionedForm,
)
from lemarche.www.tenders.tasks import (  # , send_tender_emails_to_siaes
    notify_admin_tender_created,
    send_siae_interested_email_to_author,
)
from lemarche.www.tenders.utils import create_tender_from_dict, get_or_create_user, update_or_create_questions_list


def show_sign_in_step(wizard) -> bool:
    """If the provided email matches an existing user, show sign in page"""
    # calling directly get_cleaned_data_for_step also call get_form_kwargs and leads to issues
    # because previous step data are needed in some steps
    cleaned_data = wizard.storage.get_step_data("contact")

    if cleaned_data:
        # step name 'contact' prefixes contact_email
        contact_email = cleaned_data.get("contact-contact_email")
        if contact_email and User.objects.filter(email=contact_email).exists():
            return True
    return False


class TenderCreateMultiStepView(SessionWizardView):
    """
    Multi-step Tender create form.
    Note: there is also some code in pages/views.py > csrf_failure to manage edge cases
    """

    instance = None
    success_url = reverse_lazy("siae:search_results")
    success_message = """
        <h3>Je suis Sofiane, votre conseiller chargé de vous accompagner sur votre {tender_kind}</h3>
        Votre besoin <strong>{tender_title}</strong> est bien pris en compte. <br /><br />
        Vous recevrez une notification par email dès que des prestataires seront identifiés ! <br /><br />
        À très vite
    """

    success_message_draft = """
        Votre besoin <strong>{tender_title}</strong> a été enregistré en brouillon.<br />
        Vous pourrez revenir plus tard pour le publier. Vous le retrouverez dans votre tableau de bord.
    """
    STEP_GENERAL = "general"
    STEP_DETAIL = "detail"
    STEP_CONTACT = "contact"
    STEP_SIGN_IN = "sign_in"
    STEP_SURVEY = "survey"
    STEP_CONFIRMATION = "confirmation"

    TEMPLATES = {
        STEP_GENERAL: "tenders/create_step_general.html",
        STEP_DETAIL: "tenders/create_step_detail.html",
        STEP_CONTACT: "tenders/create_step_contact.html",
        STEP_SIGN_IN: "tenders/create_step_sign_in.html",
        STEP_SURVEY: "tenders/create_step_survey.html",
        STEP_CONFIRMATION: "tenders/create_step_confirmation.html",
    }

    form_list = [
        (STEP_GENERAL, TenderCreateStepGeneralForm),
        (STEP_DETAIL, TenderCreateStepDetailForm),
        (STEP_CONTACT, TenderCreateStepContactForm),
        (STEP_SIGN_IN, TenderCreateStepSignInForm),
        (STEP_SURVEY, TenderCreateStepSurveyForm),
        (STEP_CONFIRMATION, TenderCreateStepConfirmationForm),
    ]

    # Display sign in form if the condition is True
    condition_dict = {
        "sign_in": show_sign_in_step,
    }

    # Add file storage configuration
    # https://django-formtools.readthedocs.io/en/latest/wizard.html#formtools.wizard.views.WizardView.file_storage
    file_storage = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, "temp_uploads"))

    def get_template_names(self):
        return [self.TEMPLATES[self.steps.current]]

    def get(self, request, *args, **kwargs):
        """
        Manage case when slug is sent (Tender draft edition)
        """
        if "slug" in self.kwargs:
            self.instance = get_object_or_404(Tender, slug=self.kwargs.get("slug"))
            if self.instance.status != Tender.StatusChoices.STATUS_DRAFT:
                return redirect("tenders:detail", slug=self.instance.slug)
        return super().get(request, *args, **kwargs)

    def get_form_kwargs(self, step):
        """
        Initial data
        """
        kwargs = super().get_form_kwargs(step)
        if step == self.STEP_DETAIL:
            kwargs["kind"] = self.get_cleaned_data_for_step(self.STEP_GENERAL).get("kind")
            if self.instance.id:  # for draft
                kwargs["questions_list"] = list(self.instance.questions_list())
            else:  # for new Tender, display suggested questions first
                # Do not display suggested questions for KIND_TENDER
                if kwargs["kind"] == tender_constants.KIND_TENDER:
                    kwargs["questions_list"] = []
                else:
                    kwargs["questions_list"] = [
                        {"text": question.text} for question in SuggestedQuestion.objects.all()
                    ]
        if step == self.STEP_CONTACT:
            kwargs["kind"] = self.get_cleaned_data_for_step(self.STEP_GENERAL).get("kind")
            detail_data = self.get_cleaned_data_for_step(self.STEP_DETAIL)
            kwargs["external_link"] = detail_data.get("external_link") if detail_data else None
            kwargs["user"] = self.request.user
        if step == self.STEP_SIGN_IN:
            kwargs["email"] = self.get_cleaned_data_for_step(self.STEP_CONTACT)["contact_email"]
        return kwargs

    def get_form_instance(self, step):
        """
        Manage case when slug is sent (Tender draft edition)
        """
        if "slug" in self.kwargs:
            self.instance = get_object_or_404(Tender, slug=self.kwargs.get("slug"))
        if self.instance is None:
            self.instance = Tender()
        return self.instance

    def get_context_data(self, form, **kwargs):
        """
        Pretty display of Tender at the last step
        """
        context = super().get_context_data(form=form, **kwargs)
        # needed to display the Tender preview template
        if self.steps.current == self.STEP_CONFIRMATION:
            tender_dict = self.get_all_cleaned_data()
            tender_dict["get_kind_display"] = get_choice(tender_constants.KIND_CHOICES, tender_dict["kind"])
            tender_dict["sectors_list_string"] = ", ".join(tender_dict["sectors"].values_list("name", flat=True))
            tender_dict["location_display"] = (
                Tender._meta.get_field("is_country_area").verbose_name
                if tender_dict["is_country_area"]
                else tender_dict["location"].name
            )
            tender_dict["get_amount_display"] = get_choice(
                tender_constants.AMOUNT_RANGE_CHOICES, tender_dict["amount"]
            )
            tender_dict["accept_share_amount_display"] = (
                tender_constants.ACCEPT_SHARE_AMOUNT_TRUE
                if tender_dict["accept_share_amount"]
                else tender_constants.ACCEPT_SHARE_AMOUNT_FALSE
            )
            tender_dict["attachments"] = []
            for attachment_key in ["attachment_one", "attachment_two", "attachment_three"]:
                if not tender_dict.get(f"{attachment_key}_delete"):
                    if tender_dict.get(attachment_key):
                        tender_dict["attachments"].append(tender_dict[attachment_key])
                    elif getattr(self.instance, attachment_key):
                        tender_dict["attachments"].append(getattr(self.instance, attachment_key))
            context.update({"tender": tender_dict})
        elif self.steps.current == self.STEP_DETAIL:
            tender_dict = self.get_all_cleaned_data()
            for attachment_key in ["attachment_one", "attachment_two", "attachment_three"]:
                if tender_dict.get(attachment_key):
                    context.update(
                        {
                            attachment_key: {
                                "name": os.path.basename(tender_dict.get(attachment_key).name),
                                "size": tender_dict.get(attachment_key).size,
                            }
                        }
                    )

        context["breadcrumb_links"] = []
        if self.request.user.is_authenticated:
            context["breadcrumb_links"].append(
                {"title": settings.DASHBOARD_TITLE, "url": reverse_lazy("dashboard:home")}
            )
            context["breadcrumb_links"].append({"title": "Besoins en cours", "url": reverse_lazy("tenders:list")})
        return context

    def process_step(self, form):
        """
        Save step data
        """
        data = form.data.copy()
        data["timestamp"] = timezone.now().isoformat()

        if self.request.FILES:
            data["files"] = str(self.request.FILES)

        uuid = self.request.session.get("tender_steps_data_uuid", None)
        if uuid:
            try:
                tender_steps_data = TenderStepsData.objects.get(uuid=uuid)
                tender_steps_data.steps_data.append(data)
                tender_steps_data.save()
            except TenderStepsData.DoesNotExist:
                tender_steps_data = TenderStepsData.objects.create(uuid=uuid, steps_data=[data])
        else:
            tender_steps_data = TenderStepsData.objects.create(steps_data=[data])
            self.request.session["tender_steps_data_uuid"] = tender_steps_data.uuid

        return form.data

    def save_instance_tender(self, tender_dict: dict, form_dict: dict, is_draft: bool):
        tender_status = Tender.StatusChoices.STATUS_DRAFT if is_draft else Tender.StatusChoices.STATUS_SUBMITTED
        tender_published_at = None if is_draft else timezone.now()

        if self.request.user.is_authenticated:
            tender_dict |= {
                "contact_first_name": self.request.user.first_name,
                "contact_last_name": self.request.user.last_name,
                "contact_email": self.request.user.email,
                "contact_phone": self.request.user.phone,
            }

        if self.instance.id:
            # update
            self.instance.status = tender_status
            self.instance.published_at = tender_published_at
            sectors = None
            for step, model_form in form_dict.items():
                if model_form.has_changed():
                    for attribute in model_form.changed_data:
                        match attribute:
                            case "sectors":
                                sectors = tender_dict.get("sectors", None)
                                self.instance.sectors.set(sectors)
                            case "location":
                                location = tender_dict.get("location")
                                self.instance.location = location
                                self.instance.perimeters.set([location])
                            case "questions_list":
                                update_or_create_questions_list(
                                    tender=self.instance, questions_list=tender_dict.get("questions_list")
                                )
                            case "attachment_one_delete":
                                self.instance.attachment_one = None
                            case "attachment_two_delete":
                                self.instance.attachment_two = None
                            case "attachment_three_delete":
                                self.instance.attachment_three = None
                            case _:
                                setattr(self.instance, attribute, tender_dict.get(attribute))
            # Check before adding logs or resetting modification request
            if tender_status == Tender.StatusChoices.STATUS_SUBMITTED:
                self.instance.reset_modification_request()
            self.instance.save()
        else:
            tender_dict |= {"status": tender_status, "published_at": tender_published_at}
            self.instance = create_tender_from_dict(tender_dict)

    def done(self, _, form_dict, **kwargs):
        cleaned_data = self.get_all_cleaned_data()
        # anonymous user? create user (or get an existing user by email)
        if cleaned_data.get("password"):
            user = User.objects.get(email=cleaned_data["contact_email"])
            cleaned_data.pop("password")
        else:
            user = get_or_create_user(
                self.request.user, tender_dict=cleaned_data, source=user_constants.SOURCE_TENDER_FORM
            )
        # when it's done we save the tender
        tender_instruction = TenderInstruction.objects.get(
            tender_type=cleaned_data.get("kind"), tender_source=TenderSourcesChoices.SOURCE_FORM
        )
        tender_dict = cleaned_data | {
            "author": user,
            "source": TenderSourcesChoices.SOURCE_FORM,
            "constraints": tender_instruction.text,
            "constraints_title": tender_instruction.title,
        }
        is_draft: bool = self.request.POST.get("is_draft", False)
        self.save_instance_tender(tender_dict=tender_dict, form_dict=form_dict, is_draft=is_draft)
        # remove steps data
        uuid = self.request.session.get("tender_steps_data_uuid", None)
        if uuid:
            TenderStepsData.objects.filter(uuid=uuid).delete()

        # we notify the admin team
        if settings.BITOUBI_ENV == "prod":
            notify_admin_tender_created(self.instance)

        # validation & siae contacted? in tenders/admin.py
        # success message & response
        if is_draft:
            messages.add_message(
                request=self.request,
                level=messages.INFO,
                message=self.get_success_message(cleaned_data, self.instance, is_draft=is_draft),
            )
        else:
            messages.add_message(
                request=self.request,
                level=messages.SUCCESS,
                message=self.get_success_message(cleaned_data, self.instance, is_draft=is_draft),
                extra_tags="modal_message_bizdev",
            )
        tender = self.instance
        add_to_contact_list(user=user, contact_type="signup", tender=tender)
        return redirect(self.get_success_url())

    def get_success_url(self):
        # if self.request.user.is_authenticated and not self.request.user.kind == User.KIND_SIAE:
        #     return reverse_lazy("tenders:list")  # super().get_success_url() doesn't work if called from CSRF error
        return reverse_lazy("siae:search_results")

    def get_success_message(self, cleaned_data, tender, is_draft):
        return mark_safe(
            self.success_message.format(tender_title=tender.title, tender_kind=tender.get_kind_display().lower())
            if not is_draft
            else self.success_message_draft.format(tender_title=tender.title)
        )


class TenderListView(LoginRequiredMixin, ListView):
    template_name = "tenders/list.html"
    model = Tender
    context_object_name = "tenders"
    paginate_by = 10
    paginator_class = Paginator
    status = None
    siae: Siae = None

    def get_queryset(self):
        """
        - show matching Tenders for Users KIND_SIAE
        - show owned Tenders for other Users
        """
        user = self.request.user
        qs = Tender.objects.none()
        if user.kind == User.KIND_SIAE and user.siaes:
            siaes = user.siaes.all()
            if siaes:
                # we get the first siae by default
                qs = Tender.objects.filter_with_siaes(siaes).with_is_new_for_siaes(siaes)
        else:
            qs = Tender.objects.by_user(user).with_siae_stats()
            if self.status:
                qs = qs.filter(status=self.status)

        self.filter_form = TenderFilterForm(data=self.request.GET, user=self.request.user)
        if self.filter_form.is_valid():
            kind = self.filter_form.cleaned_data.get("kind")
            if kind:
                qs = qs.filter(kind=kind)

        qs = qs.prefetch_many_to_many().select_foreign_keys()
        qs = qs.order_by_last_published()
        return qs

    def get(self, request, status=None, *args, **kwargs):
        """
        - set status
        - update 'tender_list_last_seen_date'
        """
        user = self.request.user
        self.status = status
        if user.is_authenticated:
            User.objects.filter(id=user.id).update(tender_list_last_seen_date=timezone.now())
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_kind = self.request.user.kind if self.request.user.is_authenticated else "anonymous"
        context["page_title"] = (
            settings.TENDER_DETAIL_TITLE_SIAE if user_kind == User.KIND_SIAE else settings.TENDER_DETAIL_TITLE_OTHERS
        )
        context["title_kind_sourcing_siae"] = (
            tender_constants.KIND_PROJECT_SIAE_DISPLAY
            if user_kind == User.KIND_SIAE
            else tender_constants.KIND_PROJECT_DISPLAY
        )
        context["tender_statuses"] = Tender.StatusChoices
        context["filter_form"] = self.filter_form
        context["breadcrumb_links"] = [{"title": settings.DASHBOARD_TITLE, "url": reverse_lazy("dashboard:home")}]
        return context


class TenderDetailView(TenderAuthorOrAdminRequiredIfNotSentMixin, DetailView):
    model = Tender
    template_name = "tenders/detail.html"
    context_object_name = "tender"
    is_new_for_siaes: bool = False
    object: Tender = None

    def get(self, request, *args, **kwargs):
        """
        - update 'email_link_click_date' (if ?siae_id in the URL)
        - update 'detail_display_date' (if the User has any Siae linked to this Tender)
        """
        self.object = self.get_object()
        user = self.request.user
        get_params_form = TenderDetailGetParams(request.GET)
        if get_params_form.is_valid():
            self.tender_siae = get_params_form.cleaned_data["tender_siae_uuid"]
            self.user_from_get = get_params_form.cleaned_data["user_id"]
        else:
            raise Http404()

        # update 'email_link_click_date'
        if self.tender_siae:
            if self.user_from_get:
                TenderSiae.objects.filter(
                    id=self.tender_siae.id, tender=self.object, email_link_click_date=None
                ).update(user=self.user_from_get, email_link_click_date=timezone.now(), updated_at=timezone.now())
            else:
                TenderSiae.objects.filter(
                    id=self.tender_siae.id, tender=self.object, email_link_click_date=None
                ).update(email_link_click_date=timezone.now(), updated_at=timezone.now())
        # update 'detail_display_date'
        if user.is_authenticated:
            if user.kind == User.KIND_SIAE:
                # user might not be concerned with this tender: we create TenderSiae stats
                if not user.has_tender_siae(self.object):
                    # if the user don't have the TenderSiae, the Tender is new
                    self.is_new_for_siaes = True and not self.object.deadline_date_outdated
                    for siae in user.siaes.all():
                        TenderSiae.objects.create(
                            tender=self.object, siae=siae, source=tender_constants.TENDER_SIAE_SOURCE_LINK
                        )
                # update stats
                TenderSiae.objects.filter(
                    tender=self.object, siae__in=user.siaes.all(), detail_display_date__isnull=True
                ).update(user=user, detail_display_date=timezone.now(), updated_at=timezone.now())
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        """ """
        context = super().get_context_data(**kwargs)
        # init
        user = self.request.user
        user_kind = user.kind if user.is_authenticated else "anonymous"
        show_nps = self.request.GET.get("nps", None)
        # enrich context
        context["parent_title"] = (
            settings.TENDER_DETAIL_TITLE_SIAE if user_kind == User.KIND_SIAE else settings.TENDER_DETAIL_TITLE_OTHERS
        )
        context["tender_kind_display"] = (
            tender_constants.KIND_PROJECT_SIAE_DISPLAY
            if user_kind == User.KIND_SIAE and self.object.kind == tender_constants.KIND_PROJECT
            else self.object.get_kind_display()
        )
        context["incitative_message"] = self.get_incitative_message()
        if self.tender_siae:
            context["tender_siae_uuid"] = self.tender_siae.uuid
            context["siae_has_detail_contact_click_date"] = self.tender_siae.detail_contact_click_date
            context["display_buyer_contact"] = context["siae_has_detail_contact_click_date"]
            context["siae_has_detail_not_interested_click_date"] = self.tender_siae.detail_not_interested_click_date

        context["breadcrumb_data"] = {
            "root_dir": settings_context_processors.expose_settings(self.request)["HOME_PAGE_PATH"],
            "links": [],
            "current": self.object.title[:25],
        }
        if user.is_authenticated:
            context["breadcrumb_data"]["links"].append(
                {"title": settings.DASHBOARD_TITLE, "url": reverse_lazy("dashboard:home")}
            )
            context["breadcrumb_data"]["links"].append(
                {"title": context["parent_title"], "url": reverse_lazy("tenders:list")}
            )

            if user.kind == User.KIND_SIAE:
                # Interested Siae count
                interested_count = TenderSiae.objects.filter(
                    tender=self.object, siae__in=user.siaes.all(), detail_contact_click_date__isnull=False
                ).count()
                # Hide only if all siae are already interested
                context["siae_has_detail_contact_click_date"] = (
                    interested_count
                    == TenderSiae.objects.filter(tender=self.object, siae__in=user.siaes.all()).count()
                )
                context["display_buyer_contact"] = interested_count
                if context["siae_has_detail_contact_click_date"]:
                    context["siae_has_detail_not_interested_click_date"] = False
                else:
                    context["siae_has_detail_not_interested_click_date"] = TenderSiae.objects.filter(
                        tender=self.object, siae__in=user.siaes.all(), detail_not_interested_click_date__isnull=False
                    ).exists()

                context["is_new_for_siaes"] = self.is_new_for_siaes
                if show_nps:
                    context["nps_form_id"] = settings.TALLY_SIAE_NPS_FORM_ID
            elif user.kind == User.KIND_PARTNER:
                context["user_partner_can_display_tender_contact_details"] = user.can_display_tender_contact_details
            else:
                pass
        return context

    def get_incitative_message(self):
        """Return correct wording for each kind of tender"""
        match self.object.get_kind_display():
            case tender_constants.KIND_TENDER_DISPLAY:
                message = "Soyez le premier à répondre à cet appel d'offres."
            case tender_constants.KIND_QUOTE_DISPLAY:
                message = "Soyez le premier à répondre à cette demande de devis."
            case tender_constants.KIND_PROJECT_DISPLAY:
                message = "Soyez le premier à répondre à ce projet d'achat"
            case _:
                raise NotImplementedError(f"Unknown tender kind: {self.object.kind}")
        return message


class TenderDetailContactClickStatView(SiaeUserRequiredOrTenderSiaeUUIDParamMixin, UpdateView):
    """
    Endpoint to track 'interested' button click
    We might also send a notification to the buyer
    """

    template_name = "tenders/_detail_contact_click_confirm_modal.html"
    model = Tender
    fields = []

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        self.object = self.get_object()
        self.tender_siae_uuid = request.GET.get("tender_siae_uuid", None)
        self.questions = self.object.questions.all()
        self.answers_formset_class = formset_factory(form=QuestionAnswerForm, extra=0)
        self.siae_select_form_class = SiaeSelectionForm

    def get(self, request, *args, **kwargs):
        if self.request.user.is_authenticated:
            siae_qs = Siae.objects.filter(
                users=self.request.user,
                tendersiae__tender=self.object,
                tendersiae__detail_contact_click_date__isnull=True,
            )
            siae_total_count = Siae.objects.filter(users=self.request.user, tendersiae__tender=self.object).count()

        else:  # has tender_siae_uuid
            siae_qs = Siae.objects.filter(tendersiae__uuid=self.tender_siae_uuid)
            siae_total_count = 1

        initial_data = [
            {
                "question": question,
            }
            for question in self.questions
        ]

        self.answers_formset = self.answers_formset_class(initial=initial_data)

        # Display select form only for accounts that have multiple siaes
        if siae_total_count > 1:
            self.siae_select_form = self.siae_select_form_class(
                queryset=siae_qs,
            )
        else:
            self.siae_select_form = None

        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        user = self.request.user
        self.answers_formset = self.answers_formset_class(data=self.request.POST)
        if user.is_authenticated:
            if siae_list := self.request.POST.getlist("siae"):
                siae_qs = Siae.objects.filter(id__in=siae_list)
            else:  # No siae select, mean only one matched siae
                siae_qs = Siae.objects.filter(users=self.request.user, tendersiae__tender=self.object)
        else:
            siae_qs = Siae.objects.filter(tendersiae__uuid=self.tender_siae_uuid)
        if self.answers_formset.is_valid():
            with transaction.atomic():  # Rollback all answers if any problem appears, e.g. when going back in browser
                for answer_form in self.answers_formset:
                    for siae in siae_qs:  # We copy the answer for each selected siae
                        try:  # Integrity errors can happen when going back in browser and submit again
                            QuestionAnswer.objects.create(
                                question=answer_form.cleaned_data["question"],
                                answer=answer_form.cleaned_data["answer"],
                                siae=siae,
                            )
                        except IntegrityError:
                            return self.redirect_on_error()
        else:
            return self.redirect_on_error()

        # update detail_contact_click_date
        if user.is_authenticated:
            TenderSiae.objects.filter(
                tender=self.object, siae__in=siae_qs, detail_contact_click_date__isnull=True
            ).update(user=user, detail_contact_click_date=timezone.now(), updated_at=timezone.now())
        else:
            TenderSiae.objects.filter(
                tender=self.object, siae__in=siae_qs, detail_contact_click_date__isnull=True
            ).update(detail_contact_click_date=timezone.now(), updated_at=timezone.now())

        # notify the tender author
        send_siae_interested_email_to_author(self.object)
        messages.add_message(self.request, messages.SUCCESS, self.get_success_message())

        # redirect
        return HttpResponseRedirect(self.get_success_url(self.tender_siae_uuid))

    def redirect_on_error(self):
        messages.add_message(self.request, messages.ERROR, "Une erreur a eu lieu lors de la soumission du formulaire")
        return HttpResponseRedirect(self.get_success_url(self.tender_siae_uuid))

    def get_success_url(self, tender_siae_uuid=None):
        success_url = reverse_lazy("tenders:detail", args=[self.kwargs.get("slug")])
        success_url += "?nps=true"
        if tender_siae_uuid:
            success_url += f"&tender_siae_uuid={tender_siae_uuid}"
        return success_url

    def get_success_message(self):
        return (
            "<strong>Bravo !</strong><br />"
            "Vos coordonnées, ainsi que le lien vers votre fiche commerciale ont été transmis à l'acheteur."
            " Assurez-vous d'avoir une fiche commerciale bien renseignée."
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["questions_formset"] = self.answers_formset
        if self.siae_select_form:
            ctx["siae_select_form"] = self.siae_select_form
        ctx["tender_siae_uuid"] = self.tender_siae_uuid
        return ctx


class TenderDetailNotInterestedClickView(SiaeUserRequiredOrTenderSiaeUUIDParamMixin, DetailView):
    """
    Endpoint to handle 'not interested' button click
    """

    template_name = "tenders/_detail_not_interested_click_confirm_modal.html"
    model = Tender

    def get_object(self):
        return get_object_or_404(Tender, slug=self.kwargs.get("slug"))

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        user = self.request.user
        tender_siae_uuid = request.GET.get("tender_siae_uuid", None)

        if user.is_authenticated:
            TenderSiae.objects.filter(
                tender=self.object, siae__in=user.siaes.all(), detail_not_interested_click_date__isnull=True
            ).update(
                user=user,
                detail_not_interested_feedback=self.request.POST.get("detail_not_interested_feedback", ""),
                detail_not_interested_click_date=timezone.now(),
                updated_at=timezone.now(),
            )
        else:
            TenderSiae.objects.filter(
                uuid=tender_siae_uuid, tender=self.object, detail_not_interested_click_date__isnull=True
            ).update(
                detail_not_interested_feedback=self.request.POST.get("detail_not_interested_feedback", ""),
                detail_not_interested_click_date=timezone.now(),
                updated_at=timezone.now(),
            )
        # redirect
        return HttpResponseRedirect(self.get_success_url(tender_siae_uuid))

    def get_success_url(self, tender_siae_uuid):
        success_url = reverse_lazy("tenders:detail", args=[self.kwargs.get("slug")])
        if tender_siae_uuid:
            success_url += f"?tender_siae_uuid={tender_siae_uuid}"
        return success_url


class TenderSiaeListView(TenderAuthorOrAdminRequiredMixin, FormMixin, ListView):
    template_name = "tenders/siae_interested_list.html"
    form_class = SiaeFilterForm
    queryset = Siae.objects.prefetch_related("tendersiae_set").all()
    context_object_name = "siaes"
    status = None

    def get_queryset(self):
        qs = super().get_queryset()

        # get matches references client for the current company to display a badge
        user = self.request.user
        if user.is_authenticated and user.company:
            qs = qs.with_is_company_match(company=user.company)

        # first get the tender's siaes
        self.tender = Tender.objects.get(slug=self.kwargs.get("slug"))
        qs = qs.filter_with_tender_tendersiae_status(tender=self.tender, tendersiae_status=self.status)
        # then filter with the form
        self.filter_form = SiaeFilterForm(data=self.request.GET)
        qs = self.filter_form.filter_queryset(qs)
        # Display only questions related to the current tender
        qs = qs.prefetch_related(
            Prefetch(
                "questionanswer_set",
                queryset=QuestionAnswer.objects.filter(question__tender=self.tender),
                to_attr="questions_for_tender",
            )
        )
        qs = qs.with_is_local(perimeter=self.tender.location)
        return qs

    def get(self, request, status=None, *args, **kwargs):
        """
        - set status
        - update 'siae_list_last_seen_date'
        """
        self.status = status
        Tender.objects.filter(slug=self.kwargs.get("slug")).update(siae_list_last_seen_date=timezone.now())
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["tender"] = self.tender
        context["status"] = self.status
        siae_search_form = self.filter_form if self.filter_form else SiaeFilterForm(data=self.request.GET)
        context["form"] = siae_search_form
        context["current_search_query"] = self.request.GET.urlencode()
        context["download_form"] = SiaeSelectFieldsForm(prefix="download_form")
        context["reminder_disabled"] = not self.tender.can_send_reminder
        if self.tender.deadline_date_outdated:
            context["reminder_tooltip"] = f"{self.tender.get_kind_display()} clôturé"
        elif self.tender.reminder_count > 1:
            context["reminder_tooltip"] = "Limite atteinte - Les fournisseurs ont déjà été relancés 2 fois."
        elif self.tender.reminder_last_update and (timezone.now() - self.tender.reminder_last_update) < timedelta(
            hours=24
        ):
            context["reminder_tooltip"] = (
                "Limite atteinte - Les fournisseurs ne peuvent être relancés qu'une fois en 24h"
            )
        else:
            context["reminder_tooltip"] = None

        if len(self.request.GET.keys()):
            if siae_search_form.is_valid():
                current_locations = siae_search_form.cleaned_data.get("locations")
                if current_locations:
                    context["current_locations"] = list(current_locations.values("id", "slug", "name"))

        context["breadcrumb_data"] = {
            "root_dir": settings_context_processors.expose_settings(self.request)["HOME_PAGE_PATH"],
            "links": [
                {"title": settings.DASHBOARD_TITLE, "url": reverse_lazy("dashboard:home")},
                {"title": "Mes besoins", "url": reverse_lazy("tenders:list")},
                {"title": self.tender.title[:25], "url": reverse_lazy("tenders:detail", args=[self.tender.slug])},
            ],
            "current": "Prestataires ciblés & intéressés",
        }
        return context


class TenderSiaeInterestedDownloadView(TenderAuthorOrAdminRequiredMixin, DetailView):
    http_method_names = ["get"]
    model = Tender

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        self.status = request.GET.get("tendersiae_status", None)

    def get(self, request, *args, **kwargs):
        """Gather siaes the according to the parent view (TenderSiaeListView), taking into account values from
        the filter form and the current tab (VIEWED or INTERESTED)."""
        super().get(request, *args, **kwargs)

        siae_qs = (
            SiaeFilterForm(data=self.request.GET)
            .filter_queryset(Siae.objects.filter(tendersiae__tender=self.object))
            .filter_with_tender_tendersiae_status(tender=self.object, tendersiae_status=self.status)
            .with_is_local_display(perimeter=self.object.location)
            .prefetch_related(
                Prefetch(
                    "questionanswer_set",
                    queryset=QuestionAnswer.objects.filter(question__tender=self.object).order_by("question__id"),
                    to_attr="questions_for_tender",
                )
            )
        ).order_by("name")

        form = SiaeSelectFieldsForm(data=self.request.GET, prefix="download_form")
        if form.is_valid():
            self.selected_fields = form.cleaned_data.get("selected_fields")
        else:
            return HttpResponse(400)

        if "siae_answers" in self.selected_fields:
            self.question_list = list(
                TenderQuestion.objects.filter(tender=self.object).order_by("id").values_list("text", flat=True)
            )
            self.selected_fields.remove("siae_answers")
        else:
            self.question_list = []

        header_list = self.get_selected_fields_labels(form)

        if self.request.GET.get("download_form-format") == "csv":
            return self.get_csv_response(siae_qs, header_list)
        else:
            return self.get_xlxs_response(siae_qs, header_list)

    def get_selected_fields_labels(self, form):
        """For each selected siae field in the form, return the corresponding label"""
        selected_values = form.cleaned_data["selected_fields"]

        choices_dict = dict(form.fields["selected_fields"].choices)
        selected_labels = [choices_dict[value] for value in selected_values] + self.question_list

        return selected_labels

    def get_filename(self, extension: str) -> str:
        """Get name for the exported file, according status and format."""
        if self.status == "INTERESTED":
            return f"{self.object.slug}-liste-structures-interessees.{extension}"
        elif self.status == "VIEWED":
            return f"{self.object.slug}-liste-structures-vues.{extension}"
        else:
            return f"{self.object.slug}-liste-structures-ciblees.{extension}"

    def get_csv_response(self, siae_qs, header_list):
        """Write a CSV file to a response object, containing all the defined export fields for each SIAE plus
        the questions as headers and corresponding answers in rows"""
        filename_with_extension = self.get_filename(extension="csv")

        response = HttpResponse(content_type="text/csv", charset="utf-8")
        response["Content-Disposition"] = 'attachment; filename="{}"'.format(filename_with_extension)

        writer = csv.writer(response)
        writer.writerow(header_list)

        for siae in siae_qs:
            writer.writerow(
                generate_siae_row(siae, self.selected_fields)
                + (
                    [question_answer.answer for question_answer in siae.questions_for_tender]
                    if self.question_list
                    else []
                )
            )

        return response

    def get_xlxs_response(self, siae_qs, header_list):
        """Same as get_csv_response() but for XLSX file format"""
        filename_with_extension = self.get_filename(extension="xlsx")

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="{}"'.format(filename_with_extension)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Structures"

        # header
        for index, header_item in enumerate(header_list, start=1):
            cell = ws.cell(row=1, column=index)
            cell.value = header_item
            cell.font = openpyxl.styles.Font(bold=True)

        # rows
        row_number = 2
        for siae in siae_qs:
            siae_row = generate_siae_row(siae, self.selected_fields) + (
                [question_answer.answer for question_answer in siae.questions_for_tender] if self.question_list else []
            )
            for index, row_item in enumerate(siae_row, start=1):
                cell = ws.cell(row=row_number, column=index)
                cell.value = row_item
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
            row_number += 1

        wb.save(response)

        return response


class TenderDetailSurveyTransactionedView(SesameTenderAuthorRequiredMixin, UpdateView):
    """
    Endpoint to store the tender author survey transactioned answer
    """

    template_name = "tenders/survey_transactioned_detail.html"
    form_class = TenderSurveyTransactionedForm
    queryset = Tender.objects.all()
    # success_message (see get_success_message() below)
    # success_url (see get_success_url() below)

    def get(self, request, *args, **kwargs):
        """
        Tender.survey_transactioned_answer field is updated only if:
        - the user should be the tender author (thanks to SesameTenderAuthorRequiredMixin)
        - the field is None in the database (first time answering)
        - the GET parameter 'answer' is passed
        """
        self.object = self.get_object()
        survey_transactioned_answer = request.GET.get("answer", None)
        # first time answering
        if self.object.survey_transactioned_answer in [None, constants.DONT_KNOW]:
            if survey_transactioned_answer in tender_constants.SURVEY_TRANSACTIONED_ANSWER_CHOICE_LIST:
                # update tender
                self.object.survey_transactioned_answer = survey_transactioned_answer
                self.object.survey_transactioned_answer_date = timezone.now()
                if self.object.siae_transactioned is None:
                    if survey_transactioned_answer in constants.YES_NO_CHOICE_LIST:
                        self.object.siae_transactioned = constants.YES_NO_MAPPING[survey_transactioned_answer]
                        self.object.siae_transactioned_source = (
                            tender_constants.TENDER_SIAE_TRANSACTIONED_SOURCE_AUTHOR
                        )
                self.object.save()
            else:
                pass
                # TODO or not? "answer" should always be passed
            return super().get(request, *args, **kwargs)
        # already answered
        else:
            messages.add_message(self.request, messages.WARNING, self.get_success_message(already_answered=True))
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["tender"] = self.object
        context["nps_form_id"] = settings.TALLY_BUYER_NPS_FORM_ID
        context["breadcrumb_data"] = {
            "root_dir": settings_context_processors.expose_settings(self.request)["HOME_PAGE_PATH"],
            "links": [
                {"title": settings.DASHBOARD_TITLE, "url": reverse_lazy("dashboard:home")},
                {"title": settings.TENDER_DETAIL_TITLE_OTHERS, "url": reverse_lazy("tenders:list")},
                {"title": self.object.title[:25], "url": reverse_lazy("tenders:detail", args=[self.object.slug])},
            ],
            "current": "Avez-vous contractualisé ?",
        }
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["tender_survey_transactioned_answer"] = self.object.survey_transactioned_answer
        return kwargs

    def form_valid(self, form):
        super().form_valid(form)
        messages.add_message(self.request, messages.SUCCESS, self.get_success_message())
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        success_url = reverse_lazy("tenders:detail", args=[self.kwargs.get("slug")])
        return success_url

    def get_success_message(self, already_answered=False):
        if already_answered:
            return "Votre réponse a déjà été prise en compte."
        return "Merci pour votre réponse !"


class TenderDetailSiaeSurveyTransactionedView(SesameSiaeMemberRequiredMixin, UpdateView):
    """
    Endpoint to store the tender siae survey transactioned answer
    """

    template_name = "tenders/survey_transactioned_detail.html"  # same template as author survey
    form_class = TenderSiaeSurveyTransactionedForm
    queryset = TenderSiae.objects.all()
    # success_message (see get_success_message() below)
    # success_url (see get_success_url() below)

    def get(self, request, *args, **kwargs):
        """
        TenderSiae.survey_transactioned_answer field is updated only if:
        - the user should be the tender author (thanks to SesameTenderAuthorRequiredMixin)
        - the field is None in the database (first time answering)
        - the GET parameter 'answer' is passed
        """
        self.object = self.get_object()
        survey_transactioned_answer = request.GET.get("answer", None)
        # first time answering
        if self.object.survey_transactioned_answer is None:
            if survey_transactioned_answer in ["True", "False"]:
                # transform survey_transactioned_answer into bool
                survey_transactioned_answer = survey_transactioned_answer == "True"
                # update tendersiae
                self.object.survey_transactioned_answer = survey_transactioned_answer
                self.object.survey_transactioned_answer_date = timezone.now()
                if self.object.tender.siae_transactioned is None:
                    self.object.transactioned = survey_transactioned_answer
                    self.object.transactioned_source = tender_constants.TENDER_SIAE_TRANSACTIONED_SOURCE_SIAE
                self.object.save()
                # update tender if True
                if self.object.survey_transactioned_answer:
                    if self.object.tender.siae_transactioned is None:
                        self.object.tender.siae_transactioned = survey_transactioned_answer
                        self.object.tender.siae_transactioned_source = (
                            tender_constants.TENDER_SIAE_TRANSACTIONED_SOURCE_SIAE
                        )
                        self.object.tender.save()
            else:
                pass
                # TODO or not? "answer" should always be passed
            return super().get(request, *args, **kwargs)
        # already answered
        else:
            messages.add_message(self.request, messages.WARNING, self.get_success_message(already_answered=True))
        return HttpResponseRedirect(self.get_success_url())

    def get_object(self):
        self.tender = Tender.objects.get(slug=self.kwargs.get("slug"))
        self.siae = Siae.objects.get(slug=self.kwargs.get("siae_slug"))
        return get_object_or_404(TenderSiae, tender=self.tender, siae=self.siae)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["tender"] = self.tender
        context["siae"] = self.siae
        context["breadcrumb_data"] = {
            "root_dir": settings_context_processors.expose_settings(self.request)["HOME_PAGE_PATH"],
            "links": [
                {"title": settings.DASHBOARD_TITLE, "url": reverse_lazy("dashboard:home")},
                {"title": settings.TENDER_DETAIL_TITLE_SIAE, "url": reverse_lazy("tenders:list")},
                {"title": self.tender.title[:25], "url": reverse_lazy("tenders:detail", args=[self.tender.slug])},
            ],
            "current": "Avez-vous contractualisé ?",
        }
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["tender_survey_transactioned_answer"] = self.object.survey_transactioned_answer
        return kwargs

    def form_valid(self, form):
        super().form_valid(form)
        messages.add_message(self.request, messages.SUCCESS, self.get_success_message())
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        success_url = reverse_lazy("tenders:detail", args=[self.kwargs.get("slug")])
        return success_url

    def get_success_message(self, already_answered=False):
        if already_answered:
            return "Votre réponse a déjà été prise en compte."
        return "Merci pour votre réponse !"


class TenderSiaeHideView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        tender = get_object_or_404(Tender, slug=kwargs["slug"])
        user = self.request.user
        if user.kind == User.KIND_SIAE:
            tender.tendersiae_set.filter(siae__in=user.siaes.all()).update(is_deleted_by_siae=True)
            if request.htmx:
                # status code 204 doesn't work with htmx
                return HttpResponse("", status=200)
            return HttpResponseRedirect(reverse_lazy("home"))
        else:
            # if the user is not SIAE kind, the post is not allowed
            return HttpResponse(status=401)


class TenderReminderView(TenderAuthorOrAdminRequiredMixin, SuccessMessageMixin, FormView):
    template_name = "tenders/partial_reminder_form.html"
    form_class = TenderReminderForm

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        self.tender = get_object_or_404(Tender, slug=kwargs["slug"])
        self.status = kwargs["status"]
        self.siae_qs = Siae.objects.filter_with_tender_tendersiae_status(
            tender=self.tender, tendersiae_status=self.status
        )
        # then filter with the form
        self.filter_form = SiaeFilterForm(data=self.request.GET)
        self.siae_qs = self.filter_form.filter_queryset(self.siae_qs)

    def get_success_url(self):
        return reverse_lazy("tenders:detail-siae-list", args=[self.tender.slug, self.status])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["post_url"] = reverse_lazy("tenders:send-reminder", args=[self.tender.slug, self.status])

        status_label_singular = {"VIEWED": "qui a vu"}
        status_label_plural = {"VIEWED": "qui ont vu"}
        if self.siae_qs.count() == 1:
            submit_label = f"Envoyer au fournisseur {status_label_singular.get(self.status, 'ciblé')}"
        else:
            submit_label = (
                f"Envoyer aux {self.siae_qs.count()} fournisseurs {status_label_plural.get(self.status, 'ciblés')}"
            )

        ctx["submit_button_label"] = submit_label
        return ctx

    def get_initial(self):
        initial = super().get_initial()
        initial["reminder_message"] = (
            f"Bonjour,\nplus que quelques jours pour répondre à mon besoin"
            f" “{self.tender.title}” pour {self.tender.contact_company_name}."
            f"\n\nDes questions ? Contactez moi :"
            f" {self.tender.contact_full_name} - {self.tender.contact_email} - {self.tender.contact_phone}"
        )
        return initial

    def form_valid(self, form):
        for siae in self.siae_qs:
            reversed_url = reverse_lazy("tenders:detail", kwargs={"slug": self.tender.slug})
            tender_url = f"https://{get_domain_url()}{reversed_url}"

            send_reminder_email_to_siae(
                siae,
                message=form.cleaned_data["reminder_message"],
                tender_url=tender_url,
            )
            self.tender.reminder_count += 1
            self.tender.reminder_last_update = timezone.now()
            self.tender.save()

        messages.add_message(
            self.request,
            messages.SUCCESS,
            "La relance a été envoyée avec succès à l’ensemble des fournisseurs concernés !",
        )

        return super().form_valid(form)
